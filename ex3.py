import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np

def dados_clima():

    meses             = []
    temperatura_max   = []
    temperatura_min   = []
    temperatura_media = []
    cells             = np.array([])

    book  = openpyxl.load_workbook('../Dados_climaticos_historicos.xlsx')
    sheet = book['Historico_Clima_Macae']

    for row in range(4, 8):
        for col in range(1, 14):
            char = get_column_letter(col)
            cells = np.append(cells, sheet[char + str(row)].value)

    dados = np.split(cells, 4)

    meses             = np.delete(dados[0], 0)
    temperatura_media = np.delete(dados[1], 0)
    temperatura_min   = np.delete(dados[2], 0)
    temperatura_max   = np.delete(dados[3], 0)

    print(temperatura_media)